# users/views.py

import json
import csv
from django.http import HttpResponse
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import EmailMultiAlternatives
from django.db.models import Prefetch, Q
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.generic import ListView, View, TemplateView

from tugas_akhir.models import TugasAkhir
from .forms import UserCreationAdminForm, UserEditForm
from .models import Mahasiswa, Dosen, ProgramStudi, Jurusan
from .serializers import PasswordResetConfirmSerializer, PasswordResetRequestSerializer
from django.db.models import Value
from django.db.models.functions import Concat

# openpyxl import
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Template-Based Views (Password Reset & User Management) ---

class PasswordResetRequestView(View):
    """
    Handles the initial request for a password reset.
    - Renders a form on GET.
    - On POST, validates the email, generates a reset link, and sends it.
    """
    form_template_name = 'password_reset/password_reset_request_form.html'
    success_url_name = 'users:password_reset_done'
    serializer_class = PasswordResetRequestSerializer

    def get(self, request, *args, **kwargs):
        return render(request, self.form_template_name, {'form': self.serializer_class()})

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.POST)
        success_message = 'If an account with that email exists, a password reset link has been sent.'

        if not serializer.is_valid():
            messages.success(request, success_message)
            return redirect(self.success_url_name)

        email = serializer.validated_data['email']
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.success(request, success_message)
            return redirect(self.success_url_name)

        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))

        reset_url = request.build_absolute_uri(
            reverse('users:password_reset_confirm_page', kwargs={'uidb64': uid, 'token': token})
        )

        subject = "Password Reset Request - Digita Admin"
        context = {
            'user': user,
            'reset_url': reset_url,
        }
        text_content = render_to_string('emails/password_reset.txt', context)
        html_content = render_to_string('emails/password_reset.html', context)

        try:
            email_message = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[user.email]
            )
            email_message.attach_alternative(html_content, "text/html")
            email_message.send(fail_silently=False)

        except Exception:
            messages.error(request, "There was an issue sending the email. Please try again later.")
            return render(request, self.form_template_name, {'form': serializer})

        messages.success(request, success_message)
        return redirect(self.success_url_name)


class PasswordResetConfirmView(View):
    """
    Handles the confirmation of a password reset after the user clicks the link.
    - On GET, validates the token and displays the password change form.
    - On POST, validates the new password and saves it.
    """
    form_template_name = 'password_reset/password_reset_confirm_form.html'
    invalid_token_template_name = 'password_reset/password_reset_invalid_token.html'
    success_url_name = 'users:password_reset_complete_page'
    serializer_class = PasswordResetConfirmSerializer

    def get(self, request, uidb64=None, token=None, *args, **kwargs):
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            user = None

        if user is not None and default_token_generator.check_token(user, token):
            return render(request, self.form_template_name, {'uidb64': uidb64, 'token': token})

        return render(request, self.invalid_token_template_name)

    def post(self, request, uidb64=None, token=None, *args, **kwargs):
        data = request.POST.copy()
        data.update({'uid': uidb64, 'token': token})
        serializer = self.serializer_class(data=data)

        if serializer.is_valid():
            user = serializer.validated_data['user']
            new_password = serializer.validated_data['new_password']
            user.set_password(new_password)
            user.save()
            messages.success(request, "Your password has been reset successfully. You can now log in.")
            return redirect(self.success_url_name)

        is_token_error = any(key in serializer.errors for key in ['uid', 'token'])
        if is_token_error:
            return render(request, self.invalid_token_template_name)

        context = {'uidb64': uidb64, 'token': token, 'errors_dict': serializer.errors}
        return render(request, self.form_template_name, context)


class UserManagementView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of all users and handles user creation.
    - Optimizes database queries by pre-loading related data.
    """
    model = User
    template_name = 'core/user_management.html'
    context_object_name = 'users_page'
    paginate_by = 5

    def get_queryset(self):
        queryset = User.objects.filter(
            Q(mahasiswa_profile__isnull=False) | Q(dosen_profile__isnull=False)
        ).select_related(
            'mahasiswa_profile__program_studi',
            'dosen_profile__jurusan'
        ).prefetch_related(
            'mahasiswa_profile__dosen_pembimbing__user',
            Prefetch(
                'dosen_profile__mahasiswa_binaan',
                queryset=Mahasiswa.objects.select_related('user'),
                to_attr='mahasiswa_list'
            )
        )

        search_query = self.request.GET.get('q', None)
        role_filter = self.request.GET.get('role', None)

        if search_query:
            queryset = queryset.annotate(
                full_name_search=Concat('first_name', Value(' '), 'last_name')
            ).filter(full_name_search__icontains=search_query)

        if role_filter == 'mahasiswa':
            queryset = queryset.filter(mahasiswa_profile__isnull=False)
        elif role_filter == 'dosen':
            queryset = queryset.filter(dosen_profile__isnull=False)

        return queryset.order_by('first_name', 'last_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'creation_form' not in kwargs:
            context['creation_form'] = UserCreationAdminForm()
        if 'edit_form' not in kwargs:
            context['edit_form'] = UserEditForm(instance=User())
        context['search_query'] = self.request.GET.get('q', '')
        context['role_filter'] = self.request.GET.get('role', 'all')

        paginated_users = context['users_page']
        user_list = [
            {
                'id': user.id,
                'full_name': user.get_full_name() or user.username,
                'email': user.email,
                'is_active': user.is_active,
                **self._get_user_role_data(user)
            }
            for user in paginated_users
        ]

        context['user_list'] = user_list
        context['dosen_student_json'] = json.dumps(self._get_dosen_student_map(user_list))
        context['program_studi_list'] = list(ProgramStudi.objects.values('id', 'nama_prodi'))
        context['jurusan_list'] = list(Jurusan.objects.values('id', 'nama_jurusan'))
        context['dosen_list'] = list(Dosen.objects.select_related('user').values('user_id', 'user__first_name', 'user__last_name'))
        return context

    @staticmethod
    def _get_user_role_data(user):
        if hasattr(user, 'mahasiswa_profile'):
            mahasiswa = user.mahasiswa_profile
            dospem = mahasiswa.dosen_pembimbing
            dospem_info = {
                'full_name': dospem.user.get_full_name(), 'nik': dospem.nik
            } if dospem and hasattr(dospem, 'user') else None
            return {
                'role': 'Mahasiswa',
                'identifier': mahasiswa.nim,
                'details': mahasiswa.program_studi.nama_prodi,
                'dosen_pembimbing': dospem_info,
            }
        if hasattr(user, 'dosen_profile'):
            dosen = user.dosen_profile
            mahasiswa_list = getattr(dosen, 'mahasiswa_list', [])
            mahasiswa_info = [{'full_name': m.user.get_full_name(), 'nim': m.nim} for m in mahasiswa_list]
            return {
                'role': 'Dosen',
                'identifier': dosen.nik,
                'details': dosen.jurusan.nama_jurusan,
                'dosen_user_id': user.id,
                'mahasiswa_binaan_count': len(mahasiswa_info),
                'mahasiswa_binaan_list': mahasiswa_info,
            }
        return {}

    @staticmethod
    def _get_dosen_student_map(user_list):
        return {
            u['dosen_user_id']: {'dosen_name': u['full_name'], 'students': u['mahasiswa_binaan_list']}
            for u in user_list if u.get('role') == 'Dosen'
        }

class UserExportView(LoginRequiredMixin, View):
    """
    Handles exporting the filtered user list to a polished Excel (.xlsx) file.
    """
    def get(self, request, *args, **kwargs):
        user_list_view = UserManagementView()
        user_list_view.request = self.request

        queryset = user_list_view.get_queryset().select_related(
            'mahasiswa_profile__program_studi', 'dosen_profile__jurusan'
        )

        # --- Create an in-memory Excel workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "User Data Export"

        headers = ['Full Name', 'Email', 'Role', 'NIM / NIK', 'Status', 'Details (Prodi/Jurusan)', 'Advisor / Supervised Students']
        header_font = Font(name='Calibri', bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for user in queryset:
            role_data = UserManagementView._get_user_role_data(user)
            advisor_info = ''
            if role_data.get('role') == 'Mahasiswa':
                dospem = role_data.get('dosen_pembimbing')
                advisor_info = f"{dospem['full_name']} ({dospem['nik']})" if dospem else 'Not Set'
            elif role_data.get('role') == 'Dosen':
                count = role_data.get('mahasiswa_binaan_count', 0)
                advisor_info = f"{count} Mahasiswa"

            row_data = [
                user.get_full_name() or user.username,
                user.email,
                role_data.get('role', 'N/A'),
                role_data.get('identifier', 'N/A'),
                'Active' if user.is_active else 'Inactive',
                role_data.get('details', 'N/A'),
                advisor_info
            ]
            ws.append(row_data)


        for i, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'

        wb.save(response)

        return response

class UserCreateView(LoginRequiredMixin, View):
    """
    Handles user creation via AJAX, returning JSON responses.
    """
    def post(self, request, *args, **kwargs):
        form = UserCreationAdminForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "User has been created successfully.")
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


class UserEditView(LoginRequiredMixin, View):
    """
    Handles fetching user data for editing and processing the update.
    """
    def get(self, request, pk):
        user = get_object_or_404(User.objects.select_related(
            'mahasiswa_profile__program_studi',
            'mahasiswa_profile__dosen_pembimbing__user',
            'dosen_profile__jurusan'
        ), pk=pk)

        data = {
            'id': user.id,
            'nama_lengkap': user.get_full_name(),
            'email': user.email,
            'is_active': user.is_active,
        }
        if hasattr(user, 'mahasiswa_profile'):
            profile = user.mahasiswa_profile
            data.update({
                'role': 'mahasiswa',
                'nim': profile.nim,
                'program_studi_id': profile.program_studi_id,
                'dosen_pembimbing_id': profile.dosen_pembimbing_id,
            })
        elif hasattr(user, 'dosen_profile'):
            profile = user.dosen_profile
            data.update({
                'role': 'dosen',
                'nik': profile.nik,
                'jurusan_id': profile.jurusan_id,
            })
        return JsonResponse(data)

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        form = UserEditForm(request.POST, instance=user)

        if form.is_valid():
            saved_user = form.save()
            messages.success(request, f"User '{saved_user.get_full_name()}' has been updated successfully.")

            if hasattr(saved_user, 'mahasiswa_profile'):
                profile = saved_user.mahasiswa_profile
                try:
                    tugas_akhir = TugasAkhir.objects.get(mahasiswa=profile)
                    new_dosen = profile.dosen_pembimbing

                    if tugas_akhir.dosen_pembimbing != new_dosen:
                        tugas_akhir.dosen_pembimbing = new_dosen
                        tugas_akhir.save(update_fields=['dosen_pembimbing'])
                        messages.info(request, "Advisor for the associated 'Tugas Akhir' has also been updated.")

                except TugasAkhir.DoesNotExist:
                    pass

            return redirect('users:user_management_list')

        error_message = "Update failed. " + " ".join([f"{field}: {error[0]}" for field, error in form.errors.items()])
        messages.error(request, error_message)
        return redirect('users:user_management_list')


class UserDeleteView(LoginRequiredMixin, View):
    """
    Handles the deletion of a user via a POST request.
    """
    def post(self, request, *args, **kwargs):
        user_id = kwargs.get('pk')
        try:
            user_to_delete = User.objects.get(pk=user_id)
            if user_to_delete.id == request.user.id:
                messages.error(request, "You cannot delete your own account.")
            else:
                user_full_name = user_to_delete.get_full_name() or user_to_delete.username
                user_to_delete.delete()
                messages.success(request, f"User '{user_full_name}' has been deleted successfully.")
        except User.DoesNotExist:
            messages.error(request, "The user you are trying to delete does not exist.")
        return redirect('users:user_management_list')