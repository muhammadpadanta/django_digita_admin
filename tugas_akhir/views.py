# tugas_akhir/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.views import View

import boto3
from botocore.exceptions import ClientError
from django.conf import settings
from django.http import HttpResponseRedirect, HttpResponseNotFound

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from digita_admin import settings
from users.models import Mahasiswa, Dosen
from .models import Dokumen, TugasAkhir
from .forms import DokumenEditForm, DokumenCreateForm, TugasAkhirEditForm


@login_required
def document_list_view(request):
    """
    Fetches all documents and renders them in the core/documents.html template.
    """
    search_query = request.GET.get('q', '')

    document_queryset = Dokumen.objects.select_related(
        'pemilik__user', 'pemilik__program_studi', 'tugas_akhir'
    ).order_by('-uploaded_at')

    if search_query:
        document_queryset = document_queryset.filter(
            Q(nama_dokumen__icontains=search_query) |
            Q(pemilik__user__first_name__icontains=search_query) |
            Q(pemilik__user__last_name__icontains=search_query)
        )

    # Set up Pagination
    paginator = Paginator(document_queryset, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'documents': page_obj,
        'bab_choices': Dokumen.BAB_CHOICES,
        'status_choices': Dokumen.STATUS_CHOICES,
        'tugas_akhirs': TugasAkhir.objects.all(),
        'mahasiswas': Mahasiswa.objects.select_related('user').all(),
    }

    return render(request, 'core/documents.html', context)

@login_required
def serve_document_file_view(request, pk):
    """
    Generates a pre-signed URL for an S3 object and redirects the user to it.
    This provides secure, direct-from-S3 access without proxying the file
    through the Django server.
    """
    document = get_object_or_404(Dokumen, pk=pk)

    if not document.file:
        return HttpResponseNotFound("File does not exist for this document.")

    # Get the S3 client from boto3
    s3_client = boto3.client(
        's3',
        region_name=settings.AWS_S3_REGION_NAME,
        # The following are only needed if you are not using an IAM role
        # aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        # aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
    )

    bucket_name = settings.AWS_STORAGE_BUCKET_NAME
    object_key = document.file.name

    # Set an expiration time for the URL (e.g., 15 minutes)
    expiration = 900  # Seconds

    params = {
        'Bucket': bucket_name,
        'Key': object_key,
    }

    # Check if a download is requested
    action = request.GET.get('action')
    if action == 'download':
        # To force download, we override the Content-Disposition header
        file_name = object_key.split('/')[-1]
        params['ResponseContentDisposition'] = f'attachment; filename="{file_name}"'

    try:
        # Generate the pre-signed URL
        url = s3_client.generate_presigned_url(
            'get_object',
            Params=params,
            ExpiresIn=expiration
        )
    except ClientError as e:
        # Handle potential errors (e.g., file not found on S3, permissions issue)
        print(f"Error generating pre-signed URL: {e}")
        messages.error(request, "Could not generate a secure link to the file. Please contact an administrator.")
        # Redirect back to the document list or an error page
        return redirect('tugas_akhir:document-list')

    # Redirect the user to the temporary S3 URL
    return HttpResponseRedirect(url)


class DocumentExportView(View):
    """
    Handles exporting the filtered document list to a polished Excel (.xlsx) file.
    """
    def get(self, request, *args, **kwargs):
        search_query = request.GET.get('q', '')

        # Reuse the filtering logic from the document_list_view
        document_queryset = Dokumen.objects.select_related(
            'pemilik__user', 'pemilik__program_studi', 'tugas_akhir'
        ).order_by('-uploaded_at')

        if search_query:
            document_queryset = document_queryset.filter(
                Q(nama_dokumen__icontains=search_query) |
                Q(pemilik__user__first_name__icontains=search_query) |
                Q(pemilik__user__last_name__icontains=search_query)
            )

        # --- Create an in-memory Excel workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Documents Export"

        # Define headers and apply bold styling
        headers = ['BAB', 'Nama Dokumen', 'Status', 'Pemilik', 'Program Studi', 'Judul TA', 'Waktu Upload']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        for doc in document_queryset:
            row_data = [
                doc.get_bab_display(),
                doc.nama_dokumen,
                doc.get_status_display(),
                doc.pemilik.user.get_full_name(),
                doc.pemilik.program_studi.nama_prodi,
                doc.tugas_akhir.judul or "N/A",
                doc.uploaded_at.strftime("%d %B %Y, %H:%M"),
                ]
            ws.append(row_data)

        # --- Polishing Touches ---
        # 1. Auto-adjust column widths
        for i, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 40 else 40
            ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Freeze the header row
        ws.freeze_panes = 'A2'

        # --- Prepare and return the HTTP response ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="documents_export.xlsx"'

        # Save the workbook directly to the response
        wb.save(response)

        return response

# --- VIEW for creating documents ---
@require_POST
@login_required
def create_document_view(request):
    """
    Handles the AJAX form submission for creating a new document.
    """
    print(f"DEBUG: DEFAULT_FILE_STORAGE is currently '{settings.DEFAULT_FILE_STORAGE}'")
    form = DokumenCreateForm(request.POST, request.FILES)
    if form.is_valid():
        form.save()
        messages.success(request, 'Dokumen baru telah berhasil ditambahkan.')
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)


# --- VIEW for deleting documents ---
@require_POST
@login_required
def delete_document_view(request, pk):
    """
    Handles the deletion of a document and its associated file.
    """
    document = get_object_or_404(Dokumen, pk=pk)

    if not hasattr(request.user, 'mahasiswa_profile') or document.pemilik != request.user.mahasiswa_profile:
        if not request.user.is_staff:
            messages.error(request, "Anda tidak memiliki izin untuk menghapus dokumen ini.")
            return redirect('tugas_akhir:document-list')

    try:
        document.file.delete(save=False)
        document.delete()
        messages.success(request, f"Dokumen '{document.nama_dokumen}' telah berhasil dihapus.")
    except Exception as e:
        messages.error(request, f"Terjadi kesalahan saat menghapus dokumen: {e}")

    return redirect('tugas_akhir:document-list')

@login_required
def edit_document_view(request, pk):
    document = get_object_or_404(Dokumen, pk=pk)

    if request.method == 'POST':
        form = DokumenEditForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            updated_doc = form.save()
            messages.success(request, f"Dokumen '{updated_doc.nama_dokumen}' telah berhasil diperbarui.")
            response_data = {
                'success': True,
                'document': {
                    'bab': updated_doc.get_bab_display(),
                    'nama_dokumen': updated_doc.nama_dokumen,
                    'status': updated_doc.get_status_display(),
                    'pemilik_name': updated_doc.pemilik.user.get_full_name(),
                    'pemilik_prodi': updated_doc.pemilik.program_studi.nama_prodi,
                    'file_url': updated_doc.file.url if updated_doc.file else '#',
                }
            }
            return JsonResponse(response_data)
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    else:
        data = {
            'tugas_akhir': document.tugas_akhir_id,
            'bab': document.bab,
            'nama_dokumen': document.nama_dokumen,
            'status': document.status,
            'pemilik': document.pemilik_id,
            'current_file_url': document.file.url if document.file else None,
            'current_file_name': document.file.name.split('/')[-1] if document.file else 'No file uploaded',
        }
        return JsonResponse(data)


def tugas_akhir_list_view(request):
    """
    Fetches all Tugas Akhir objects and renders them in the ta_list.html template.
    """
    search_query = request.GET.get('q', '')

    tugas_akhir_queryset = TugasAkhir.objects.select_related(
        'mahasiswa__user',
        'dosen_pembimbing__user'
    ).order_by('-created_at')

    if search_query:
        tugas_akhir_queryset = tugas_akhir_queryset.filter(
            Q(judul__icontains=search_query) |
            Q(deskripsi__icontains=search_query) |
            Q(mahasiswa__user__first_name__icontains=search_query) |
            Q(mahasiswa__user__last_name__icontains=search_query) |
            Q(dosen_pembimbing__user__first_name__icontains=search_query) |
            Q(dosen_pembimbing__user__last_name__icontains=search_query)
        )

    # Set up Pagination (optional, but good practice)
    paginator = Paginator(tugas_akhir_queryset, 3)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tugas_akhir_list': page_obj,
        'search_query': search_query,
    }
    return render(request, 'core/tugas_akhir.html', context)

class TugasAkhirExportView(View):
    """
    Handles exporting the filtered Tugas Akhir list to a polished Excel (.xlsx) file.
    """
    def get(self, request, *args, **kwargs):
        search_query = request.GET.get('q', '')

        tugas_akhir_queryset = TugasAkhir.objects.select_related(
            'mahasiswa__user',
            'dosen_pembimbing__user'
        ).order_by('-created_at')

        if search_query:
            tugas_akhir_queryset = tugas_akhir_queryset.filter(
                Q(judul__icontains=search_query) |
                Q(deskripsi__icontains=search_query) |
                Q(mahasiswa__user__first_name__icontains=search_query) |
                Q(mahasiswa__user__last_name__icontains=search_query) |
                Q(dosen_pembimbing__user__first_name__icontains=search_query) |
                Q(dosen_pembimbing__user__last_name__icontains=search_query)
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Tugas Akhir Export"

        headers = ['ID', 'Judul', 'Deskripsi', 'Dosen Pembimbing', 'Mahasiswa', 'NIM', 'Tanggal Dibuat']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for ta in tugas_akhir_queryset:
            row_data = [
                f"TA{ta.pk:03d}",
                ta.judul or "-",
                ta.deskripsi or "-",
                ta.dosen_pembimbing.user.get_full_name() if ta.dosen_pembimbing and ta.dosen_pembimbing.user else "Belum Ditentukan",
                ta.mahasiswa.user.get_full_name() if ta.mahasiswa and ta.mahasiswa.user else "N/A",
                ta.mahasiswa.nim if ta.mahasiswa else "N/A",
                ta.created_at.strftime("%d %B %Y, %H:%M"),
                ]
            ws.append(row_data)

        for i, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="tugas_akhir_export.xlsx"'

        wb.save(response)

        return response

def tugas_akhir_detail_view(request, pk):
    """
    Returns the details of a single TugasAkhir instance as JSON.
    """
    try:
        ta = TugasAkhir.objects.select_related(
            'mahasiswa__user',
            'dosen_pembimbing__user'
        ).get(pk=pk)

        data = {
            'id': f"TA{ta.pk:03d}",
            'judul': ta.judul or "Belum ada judul",
            'deskripsi': ta.deskripsi or "Tidak ada deskripsi.",
            'mahasiswa': ta.mahasiswa.user.get_full_name() if ta.mahasiswa and ta.mahasiswa.user else "N/A",
            'nim': ta.mahasiswa.nim if ta.mahasiswa else "N/A",
            'dosen_pembimbing': ta.dosen_pembimbing.user.get_full_name() if ta.dosen_pembimbing and ta.dosen_pembimbing.user else "Belum Ditentukan",
            'created_at': ta.created_at.strftime("%d %B %Y, %H:%M"),
        }
        return JsonResponse(data)
    except TugasAkhir.DoesNotExist:
        return JsonResponse({'error': 'Tugas Akhir not found'}, status=404)


@require_POST
@login_required
def delete_tugas_akhir_view(request, pk):
    """
    Handles the deletion of a TugasAkhir instance.
    """
    tugas_akhir = get_object_or_404(TugasAkhir, pk=pk)

    try:
        ta_title = tugas_akhir.judul or f"TA Mahasiswa {tugas_akhir.mahasiswa.user.get_full_name()}"
        tugas_akhir.delete()
        messages.success(request, f"Tugas Akhir '{ta_title}' telah berhasil dihapus.")
    except Exception as e:
        messages.error(request, f"Terjadi kesalahan saat menghapus Tugas Akhir: {e}")

    return redirect('tugas_akhir:ta-list')

@login_required
def edit_tugas_akhir_view(request, pk):
    """
    Handles fetching and updating a TugasAkhir instance via AJAX.
    """
    tugas_akhir = get_object_or_404(TugasAkhir, pk=pk)

    if request.method == 'POST':
        form = TugasAkhirEditForm(request.POST, instance=tugas_akhir)
        if form.is_valid():
            updated_ta = form.save()
            dosen_name = "Belum Ditentukan"
            if updated_ta.dosen_pembimbing:
                dosen_name = updated_ta.dosen_pembimbing.user.get_full_name()

            messages.success(request, f"Tugas Akhir '{updated_ta.judul}' telah berhasil diperbarui.")

            return JsonResponse({
                'success': True,
                'tugas_akhir': {
                    'id': updated_ta.pk,
                    'judul': updated_ta.judul or "-",
                    'deskripsi': updated_ta.deskripsi,
                    'dosen_pembimbing_name': dosen_name,
                }
            })
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)

    # Handle GET request
    else:
        dosens = list(Dosen.objects.select_related('user').values('pk', 'user__first_name', 'user__last_name'))
        data = {
            'id': tugas_akhir.pk,
            'judul': tugas_akhir.judul,
            'deskripsi': tugas_akhir.deskripsi,
            'dosen_pembimbing_id': tugas_akhir.dosen_pembimbing_id,
            'dosens': dosens,
        }
        return JsonResponse(data)

def get_tugas_akhir_for_mahasiswa(request, mahasiswa_id):
    """
    API endpoint to fetch the TugasAkhir for a given Mahasiswa ID.
    Returns the TugasAkhir's ID and string representation as JSON.
    """
    try:
        tugas_akhir = TugasAkhir.objects.get(mahasiswa_id=mahasiswa_id)
        data = {
            'success': True,
            'tugas_akhir_id': tugas_akhir.pk,
            'tugas_akhir_str': str(tugas_akhir)
        }
    except TugasAkhir.DoesNotExist:
        data = {
            'success': False,
            'error': 'Mahasiswa ini belum memiliki data Tugas Akhir.'
        }
    return JsonResponse(data)

