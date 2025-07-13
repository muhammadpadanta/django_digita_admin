# tugas_akhir/tests.py
import io
import pytest
from unittest.mock import patch, MagicMock
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db.utils import IntegrityError
from django.urls import reverse
from openpyxl import load_workbook

from .models import TugasAkhir, RequestDosen, Dokumen, JadwalBimbingan, Ruangan
from users.models import Mahasiswa, Dosen, Jurusan, ProgramStudi
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile

pytestmark = pytest.mark.django_db

# --- Fixtures ---

@pytest.fixture
def jurusan():
    return Jurusan.objects.create(nama_jurusan='Teknik Elektro')

@pytest.fixture
def prodi(jurusan):
    return ProgramStudi.objects.create(nama_prodi='D3 Teknik Listrik', jurusan=jurusan)

@pytest.fixture
def admin_user():
    return User.objects.create_superuser(username='admin', password='password', email='admin@example.com')

@pytest.fixture
def dosen_user():
    return User.objects.create_user(username='dosen1', password='password', first_name='Andi', last_name='Wijaya')

@pytest.fixture
def dosen(dosen_user, jurusan):
    return Dosen.objects.create(user=dosen_user, nik='12345', jurusan=jurusan)

@pytest.fixture
def mahasiswa_user():
    return User.objects.create_user(username='mahasiswa1', password='password', first_name='Budi', last_name='Santoso')

@pytest.fixture
def mahasiswa(mahasiswa_user, prodi):
    return Mahasiswa.objects.create(user=mahasiswa_user, nim='54321', program_studi=prodi)

@pytest.fixture
def tugas_akhir(mahasiswa, dosen):
    return TugasAkhir.objects.create(
        mahasiswa=mahasiswa,
        dosen_pembimbing=dosen,
        judul="Rancang Bangun Sistem Informasi"
    )

@pytest.fixture
def ruangan():
    return Ruangan.objects.create(nama_ruangan="Ruang Rapat 1", gedung="Gedung A")

@pytest.fixture
def pdf_file():
    return SimpleUploadedFile("test.pdf", b"file_content", content_type="application/pdf")

@pytest.fixture
def document(tugas_akhir, mahasiswa, pdf_file):
    # Mock the S3 storage
    with patch('tugas_akhir.models.S3Boto3Storage'):
        doc = Dokumen.objects.create(
            tugas_akhir=tugas_akhir,
            pemilik=mahasiswa,
            bab='BAB I',
            nama_dokumen='Pendahuluan.pdf',
            file=pdf_file
        )
        return doc

# --- Model Tests ---

class TestTugasAkhirModel:
    def test_tugas_akhir_creation(self, tugas_akhir, mahasiswa, dosen):
        assert tugas_akhir.mahasiswa == mahasiswa
        assert tugas_akhir.dosen_pembimbing == dosen
        assert str(tugas_akhir) == f"TA {mahasiswa.nim} - {tugas_akhir.judul}"

class TestRequestDosenModel:
    def test_request_dosen_creation(self, mahasiswa, dosen):
        request = RequestDosen.objects.create(
            mahasiswa=mahasiswa,
            dosen=dosen,
            rencana_judul="Judul Awal",
            rencana_deskripsi="Deskripsi Awal"
        )
        assert request.status == 'PENDING'
        assert str(request) == f"Request from {mahasiswa.nim} to {dosen.nik} (PENDING)"

    def test_unique_pending_request_constraint(self, mahasiswa, dosen):
        RequestDosen.objects.create(
            mahasiswa=mahasiswa,
            dosen=dosen,
            status='PENDING',
            rencana_judul="Judul 1",
            rencana_deskripsi="Deskripsi 1"
        )
        with pytest.raises(IntegrityError):
            RequestDosen.objects.create(
                mahasiswa=mahasiswa,
                dosen=dosen,
                status='PENDING',
                rencana_judul="Judul 2",
                rencana_deskripsi="Deskripsi 2"
            )

class TestDokumenModel:
    def test_dokumen_creation(self, tugas_akhir, mahasiswa):
        dokumen = Dokumen.objects.create(
            tugas_akhir=tugas_akhir,
            pemilik=mahasiswa,
            bab='BAB I',
            nama_dokumen='Pendahuluan.pdf'
        )
        assert dokumen.status == 'Pending'
        assert str(dokumen) == f"{dokumen.nama_dokumen} ({dokumen.get_bab_display()}) - {mahasiswa.user.get_full_name()}"

    def test_unique_bab_per_mahasiswa_constraint(self, tugas_akhir, mahasiswa):
        Dokumen.objects.create(
            tugas_akhir=tugas_akhir,
            pemilik=mahasiswa,
            bab='BAB I',
            nama_dokumen='Pendahuluan_v1.pdf'
        )
        with pytest.raises(IntegrityError):
            Dokumen.objects.create(
                tugas_akhir=tugas_akhir,
                pemilik=mahasiswa,
                bab='BAB I',
                nama_dokumen='Pendahuluan_v2.pdf'
            )

class TestJadwalBimbinganModel:
    def test_jadwal_bimbingan_creation(self, mahasiswa, dosen, ruangan):
        jadwal = JadwalBimbingan.objects.create(
            mahasiswa=mahasiswa,
            dosen_pembimbing=dosen,
            judul_bimbingan="Diskusi Bab 1",
            tanggal=timezone.now().date(),
            waktu=timezone.now().time(),
            lokasi_ruangan=ruangan,
            lokasi_text="-"
        )
        assert jadwal.status == 'PENDING'
        assert str(jadwal) == f"Bimbingan {mahasiswa.nim} - {jadwal.tanggal} ({jadwal.get_status_display()})"

class TestRuanganModel:
    def test_ruangan_creation(self):
        ruangan = Ruangan.objects.create(nama_ruangan="Lab Komputer", gedung="FTI")
        assert str(ruangan) == "Lab Komputer (FTI)"


# --- View Tests ---

class TestDocumentListView:
    def test_document_list_view_get(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-list')
        response = client.get(url)
        assert response.status_code == 200
        assert 'core/documents.html' in [t.name for t in response.templates]
        assert document in response.context['documents']

    def test_document_list_view_search(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-list')
        response = client.get(url, {'q': 'Pendahuluan'})
        assert response.status_code == 200
        assert document in response.context['documents']
        response = client.get(url, {'q': 'NonExistent'})
        assert response.status_code == 200
        assert document not in response.context['documents']

@patch('tugas_akhir.views.boto3.client')
def test_serve_document_file_view(mock_boto3_client, client, admin_user, document):
    mock_s3 = MagicMock()
    mock_s3.generate_presigned_url.return_value = 'https://s3.amazonaws.com/some/presigned/url'
    mock_boto3_client.return_value = mock_s3

    client.force_login(admin_user)
    url = reverse('tugas_akhir:document-file', kwargs={'pk': document.pk})
    response = client.get(url)

    assert response.status_code == 302
    assert response.url == 'https://s3.amazonaws.com/some/presigned/url'

class TestDocumentExportView:
    def test_export_view(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-export')
        response = client.get(url)

        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Check content
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet['A1'].value == 'BAB'
        assert sheet['B2'].value == document.nama_dokumen

class TestDocumentCreateView:
    def test_create_document_success(self, client, admin_user, tugas_akhir, mahasiswa, pdf_file):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-create')
        data = {
            'tugas_akhir': tugas_akhir.pk,
            'pemilik': mahasiswa.pk,
            'bab': 'BAB II',
            'nama_dokumen': 'Test Doc',
            'file': pdf_file,
        }
        response = client.post(url, data)
        assert response.status_code == 200
        assert response.json()['success'] is True
        assert Dokumen.objects.filter(nama_dokumen='Test Doc').exists()

    def test_create_document_fail(self, client, admin_user):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-create')
        response = client.post(url, {}) # Invalid data
        assert response.status_code == 400
        assert response.json()['success'] is False

class TestDocumentDeleteView:
    def test_delete_document_staff(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-delete', kwargs={'pk': document.pk})
        response = client.post(url)
        assert response.status_code == 302 # Redirect
        assert not Dokumen.objects.filter(pk=document.pk).exists()

    def test_delete_document_owner(self, client, mahasiswa_user, document):
        document.pemilik = mahasiswa_user.mahasiswa_profile
        document.save()
        client.force_login(mahasiswa_user)
        url = reverse('tugas_akhir:document-delete', kwargs={'pk': document.pk})
        response = client.post(url)
        assert response.status_code == 302
        assert not Dokumen.objects.filter(pk=document.pk).exists()

class TestDocumentEditView:
    def test_edit_document_get(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-edit', kwargs={'pk': document.pk})
        response = client.get(url)
        assert response.status_code == 200
        assert response.json()['nama_dokumen'] == document.nama_dokumen

    def test_edit_document_post_success(self, client, admin_user, document):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:document-edit', kwargs={'pk': document.pk})
        data = {
            'tugas_akhir': document.tugas_akhir.pk,
            'pemilik': document.pemilik.pk,
            'bab': document.bab,
            'nama_dokumen': 'Updated Name',
            'status': 'Disetujui',
        }
        response = client.post(url, data)
        assert response.status_code == 200
        assert response.json()['success'] is True
        document.refresh_from_db()
        assert document.nama_dokumen == 'Updated Name'

class TestTugasAkhirListView:
    def test_tugas_akhir_list_view(self, client, admin_user, tugas_akhir):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-list')
        response = client.get(url)
        assert response.status_code == 200
        assert 'core/tugas_akhir.html' in [t.name for t in response.templates]
        assert tugas_akhir in response.context['tugas_akhir_list']

class TestTugasAkhirExportView:
    def test_export_view(self, client, admin_user, tugas_akhir):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-export')
        response = client.get(url)
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet['B2'].value == tugas_akhir.judul

class TestTugasAkhirDetailView:
    def test_detail_view_success(self, client, admin_user, tugas_akhir):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-detail', kwargs={'pk': tugas_akhir.pk})
        response = client.get(url)
        assert response.status_code == 200
        assert response.json()['judul'] == tugas_akhir.judul

    def test_detail_view_fail(self, client, admin_user):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-detail', kwargs={'pk': 999})
        response = client.get(url)
        assert response.status_code == 404

class TestTugasAkhirDeleteView:
    def test_delete_view(self, client, admin_user, tugas_akhir):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-delete', kwargs={'pk': tugas_akhir.pk})
        response = client.post(url)
        assert response.status_code == 302
        assert not TugasAkhir.objects.filter(pk=tugas_akhir.pk).exists()

class TestTugasAkhirEditView:
    def test_edit_view_get(self, client, admin_user, tugas_akhir):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-edit', kwargs={'pk': tugas_akhir.pk})
        response = client.get(url)
        assert response.status_code == 200
        assert response.json()['judul'] == tugas_akhir.judul

    def test_edit_view_post(self, client, admin_user, tugas_akhir, dosen):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:ta-edit', kwargs={'pk': tugas_akhir.pk})
        data = {
            'judul': 'New Judul',
            'deskripsi': 'New Deskripsi',
            'dosen_pembimbing': dosen.pk,
        }
        response = client.post(url, data)
        assert response.status_code == 200
        assert response.json()['success'] is True
        tugas_akhir.refresh_from_db()
        assert tugas_akhir.judul == 'New Judul'

class TestGetTugasAkhirForMahasiswaView:
    def test_get_tugas_akhir_success(self, client, admin_user, tugas_akhir, mahasiswa):
        client.force_login(admin_user)
        url = reverse('tugas_akhir:api-get-ta-for-mahasiswa', kwargs={'mahasiswa_id': mahasiswa.pk})
        response = client.get(url)
        assert response.status_code == 200
        assert response.json()['success'] is True
        assert response.json()['tugas_akhir_id'] == tugas_akhir.pk

    def test_get_tugas_akhir_fail(self, client, admin_user, mahasiswa):
        # Log in the user
        client.force_login(admin_user)

        # Create a new mahasiswa without a tugas_akhir
        user = User.objects.create_user(username='new_mahasiswa', password='password')
        new_mahasiswa = Mahasiswa.objects.create(user=user, nim='67890', program_studi=mahasiswa.program_studi)

        # Make the request
        url = reverse('tugas_akhir:api-get-ta-for-mahasiswa', kwargs={'mahasiswa_id': new_mahasiswa.pk})
        response = client.get(url)

        # Assert the outcome
        assert response.status_code == 200
        assert response.json()['success'] is False
        assert 'error' in response.json()